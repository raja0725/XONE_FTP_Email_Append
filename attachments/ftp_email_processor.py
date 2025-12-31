"""
FTP Email Processing Script
Handles file download, conversion, upload, and email append reporting
Django-ready implementation
"""
import paramiko
import os
import csv
import ftplib
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import requests
import json

SKIP_FTP1_DOWNLOAD = False
SKIP_EMAIL_SEND = True
LOCAL_TEST_FILE = r"C:\Users\Admin\OneDrive\Documents\FPN_email_append\Test_data\2025-12-09_Fushia_Upload_FPN_TEST.csv"

class PixlProcessor:
    def __init__(self, config):
        self.config = config

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('ftp_processor.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class FTPEmailProcessor:
    """Main processor for FTP file operations and email append reporting"""
    def connect_sftp(self, server: str, username: str, password: str, port: int = 22):
        logger.info(f"Connecting to SFTP server: {server}")
        try:
            # Use SSHClient for better authentication handling
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(
                hostname=server,
                port=port,
                username=username,
                password=password,
                look_for_keys=False,  # Don't look for SSH keys, use password only
                allow_agent=False,    # Don't use SSH agent
                timeout=30
            )
            sftp = ssh.open_sftp()
            logger.info("SFTP connection successful")
            return sftp, ssh
        except paramiko.AuthenticationException as e:
            logger.error(f"Authentication failed: {e}")
            logger.error("Please verify:")
            logger.error("  1. Username and password are correct in .env file")
            logger.error("  2. TowerData account is active")
            logger.error("  3. Your IP is whitelisted (if required)")
            raise
        except Exception as e:
            logger.error(f"Connection failed: {e}")
            raise

    def upload_to_sftp(self, sftp: paramiko.SFTPClient, local_filename: str, remote_filename: str):
        logger.info(f"Uploading {local_filename} to SFTP as {remote_filename}")
        sftp.put(local_filename, remote_filename)
        logger.info("File uploaded successfully (SFTP)")

    def download_from_sftp(self, sftp: paramiko.SFTPClient, remote_filename: str, local_filename: str) -> str:
        logger.info(f"Downloading {remote_filename} from SFTP")
        local_path = self.download_dir / local_filename
        sftp.get(remote_filename, str(local_path))
        logger.info(f"File downloaded successfully to {local_path}")
        return str(local_path)

    def sftp_list_files(self, sftp: paramiko.SFTPClient):
        return [f.filename for f in sftp.listdir_attr()]

    def wait_for_processing_sftp(self, sftp: paramiko.SFTPClient, upload_filename: str,
                                 max_wait_minutes: int = 30, check_interval: int = 60):
        """
        Wait for SFTP to process file and make it available in downloads/ directory
        TowerData puts processed files in downloads/ with pattern: filename-results-XXXXX.csv

        Args:
            sftp: SFTP connection object
            upload_filename: Name of uploaded file (used to match results)
            max_wait_minutes: Maximum time to wait in minutes
            check_interval: Time between checks in seconds

        Returns:
            Filename of the processed file in downloads/ directory
        """
        # Remove .csv extension and create search pattern
        base_name = upload_filename.replace('.csv', '')
        # TowerData removes dots from filenames, so normalize for pattern matching
        base_name_normalized = base_name.replace('.', '')
        logger.info(f"Waiting for processing to complete.")
        logger.info(f"Looking for files matching: {base_name_normalized}*-results-*.csv in downloads/ directory")

        max_attempts = (max_wait_minutes * 60) // check_interval
        attempts = 0

        while attempts < max_attempts:
            try:
                # List files in downloads directory
                files = sftp.listdir_attr('downloads')

                # Search for matching file
                for file_attr in files:
                    filename = file_attr.filename
                    # Check if file matches our pattern and was modified recently (within last hour)
                    # Use normalized base_name (without dots) since TowerData removes them
                    if (base_name_normalized in filename and '-results-' in filename and
                        filename.endswith('.csv')):
                        # Check if file was modified recently (within the last 2 hours to catch it)
                        import time as time_module
                        file_age_hours = (time_module.time() - file_attr.st_mtime) / 3600

                        if file_age_hours < 2:  # File modified within last 2 hours
                            logger.info(f"[OK] Found processed file: {filename}")
                            logger.info(f"  Modified {file_age_hours:.1f} hours ago")
                            return f"downloads/{filename}"

                attempts += 1
                logger.info(f"File not ready yet. Attempt {attempts}/{max_attempts}. Waiting {check_interval} seconds...")
                time.sleep(check_interval)

            except Exception as e:
                logger.warning(f"Error checking for file: {str(e)}")
                attempts += 1
                time.sleep(check_interval)

        logger.error(f"Processed file not found after {max_wait_minutes} minutes")
        logger.error(f"Expected pattern: {base_name_normalized}*-results-*.csv in downloads/ directory")
        raise TimeoutError(f"Processing timeout: No matching file found in downloads/")

    def __init__(self, config: Dict):
        """
        Initialize processor with configuration

        Args:
            config: Dictionary containing FTP credentials and settings
        """
        self.config = config
        self.download_dir = Path(config.get('download_dir', './downloads'))
        self.output_dir = Path(config.get('output_dir', './output'))
        self.download_dir.mkdir(parents=True, exist_ok=True)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # File tracking
        self.original_file = None
        self.converted_file = None
        self.appended_file = None

        # Email statistics
        self.emails_before = 0
        self.emails_after = 0
        self.new_emails = []
        self.duplicate_emails = []

    def connect_ftp(self, server: str, username: str, password: str, port: int = 21) -> ftplib.FTP:
        """
        Establish FTP connection with TLS/SSL support

        Args:
            server: FTP server address
            username: FTP username
            password: FTP password
            port: FTP port (default 21)

        Returns:
            FTP connection object
        """
        try:
            logger.info(f"Connecting to FTP server: {server} (with TLS/SSL)")
            # Use FTP_TLS for secure connection
            ftp = ftplib.FTP_TLS()
            ftp.connect(server, port)
            ftp.login(username, password)
            # Secure the data connection
            ftp.prot_p()
            logger.info("FTP connection successful (TLS/SSL enabled)")
            return ftp
        except Exception as e:
            logger.error(f"FTP connection failed: {str(e)}")
            raise

    def download_from_ftp(self, ftp: ftplib.FTP, remote_filename: str, local_filename: str) -> str:
        """
        Download file from FTP server

        Args:
            ftp: FTP connection object
            remote_filename: Name of file on FTP server
            local_filename: Local path to save file

        Returns:
            Path to downloaded file
        """
        try:
            logger.info(f"Downloading {remote_filename} from FTP server")
            local_path = self.download_dir / local_filename

            with open(local_path, 'wb') as f:
                ftp.retrbinary(f'RETR {remote_filename}', f.write)

            logger.info(f"File downloaded successfully to {local_path}")
            return str(local_path)
        except Exception as e:
            logger.error(f"Download failed: {str(e)}")
            raise

    def upload_to_ftp(self, ftp: ftplib.FTP, local_filename: str, remote_filename: str):
        """
        Upload file to FTP server

        Args:
            ftp: FTP connection object
            local_filename: Local file path
            remote_filename: Remote filename to save as
        """
        try:
            logger.info(f"Uploading {local_filename} to FTP server as {remote_filename}")

            with open(local_filename, 'rb') as f:
                ftp.storbinary(f'STOR {remote_filename}', f)

            logger.info("File uploaded successfully")
        except Exception as e:
            logger.error(f"Upload failed: {str(e)}")
            raise

    def list_ftp_files(self, ftp: ftplib.FTP, directory: str, keyword: str = None) -> List[str]:
        """
        List files in FTP directory with optional keyword filter

        Args:
            ftp: FTP connection object
            directory: Directory path on FTP server
            keyword: Optional keyword to filter files (case-insensitive)

        Returns:
            List of matching filenames
        """
        try:
            logger.info(f"Listing files in {directory}")
            ftp.cwd(directory)
            all_files = ftp.nlst()

            if keyword:
                filtered_files = [f for f in all_files if keyword.lower() in f.lower()]
                logger.info(f"Found {len(filtered_files)} files containing '{keyword}'")
                return filtered_files

            logger.info(f"Found {len(all_files)} files")
            return all_files
        except Exception as e:
            logger.error(f"Failed to list files in {directory}: {str(e)}")
            raise

    def ensure_ftp_directory(self, ftp: ftplib.FTP, directory: str):
        """
        Ensure a directory exists on FTP server, create if it doesn't

        Args:
            ftp: FTP connection object
            directory: Directory path to ensure exists
        """
        # Normalize path - remove trailing slash for consistency
        directory = directory.rstrip('/')

        try:
            # Try to change to the directory
            ftp.cwd(directory)
            logger.info(f"Directory exists: {directory}")
            # Return to original directory
            ftp.cwd(self.config['source_filename'].rstrip('/'))
        except:
            # Directory doesn't exist, create it
            try:
                logger.info(f"Creating directory: {directory}")
                ftp.mkd(directory)
                logger.info(f"Directory created: {directory}")
            except Exception as e:
                logger.warning(f"Could not create directory {directory}: {str(e)}")

    def move_ftp_file(self, ftp: ftplib.FTP, source_path: str, dest_path: str):
        """
        Move file on FTP server (rename to different directory)

        Args:
            ftp: FTP connection object
            source_path: Current file path
            dest_path: Destination file path
        """
        try:
            logger.info(f"Moving {source_path} to {dest_path}")
            ftp.rename(source_path, dest_path)
            logger.info("File moved successfully")
        except Exception as e:
            logger.error(f"Failed to move file: {str(e)}")
            raise

    def send_email_notification(self, subject: str, body: str, html: bool = True):
        """
        Send email notification via SMTP

        Args:
            subject: Email subject
            body: Email body (HTML or plain text)
            html: Whether body is HTML (default True)
        """
        if not self.config.get('send_email_notification', False):
            logger.info("Email notifications disabled")
            return

        try:
            email_from = self.config.get('email_from')
            email_to = self.config.get('email_to')
            smtp_server = self.config.get('smtp_server')
            smtp_port = self.config.get('smtp_port', 587)
            smtp_username = self.config.get('smtp_username')
            smtp_password = self.config.get('smtp_password')

            if not all([email_from, email_to, smtp_server, smtp_username, smtp_password]):
                logger.warning("Email configuration incomplete, skipping notification")
                return

            # Create message
            msg = MIMEMultipart('alternative')
            msg['From'] = email_from
            msg['To'] = email_to
            msg['Subject'] = subject

            # Attach body
            if html:
                msg.attach(MIMEText(body, 'html'))
            else:
                msg.attach(MIMEText(body, 'plain'))

            # Send email
            logger.info(f"Sending email to {email_to}...")
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(smtp_username, smtp_password)
                server.send_message(msg)

            logger.info("[OK] Email notification sent successfully")

        except Exception as e:
            logger.error(f"Failed to send email notification: {str(e)}")

    def send_slack_notification(self, message: str, title: str = None, mention_channel: bool = True):
        """
        Send notification to Slack channel via webhook

        Args:
            message: Message text
            title: Optional title/heading
            mention_channel: If True, adds @channel mention to notify all members (default: True)
        """
        if not self.config.get('send_slack_notification', False):
            logger.info("Slack notifications disabled")
            return

        try:
            webhook_url = self.config.get('slack_webhook_url')
            channel = self.config.get('slack_channel', '#general')

            if not webhook_url or webhook_url == 'https://hooks.slack.com/services/YOUR/WEBHOOK/URL':
                logger.warning("Slack webhook URL not configured, skipping notification")
                return

            # Add @channel mention for visibility if requested
            if mention_channel:
                message = f"<!channel>\n{message}"

            # Build Slack message payload
            payload = {
                "channel": channel,
                "username": "FPN Email Processor",
                "icon_emoji": ":email:",
                "link_names": 1,  # Enable @mentions
            }

            if title:
                payload["attachments"] = [{
                    "color": "#36a64f",
                    "title": title,
                    "text": message,
                    "footer": "FPN Email Append Processor",
                    "ts": int(datetime.now().timestamp())
                }]
            else:
                payload["text"] = message

            # Send to Slack
            logger.info(f"Sending Slack notification to {channel}...")
            response = requests.post(webhook_url, json=payload, timeout=10)

            if response.status_code == 200:
                logger.info("[OK] Slack notification sent successfully")
            else:
                logger.warning(f"Slack notification failed: {response.status_code} - {response.text}")

        except Exception as e:
            logger.error(f"Failed to send Slack notification: {str(e)}")

    def convert_pipe_to_csv(self, input_file: str, output_file: str = None) -> str:
        """
        Convert pipe-delimited file to comma-delimited CSV with proper quoting

        Args:
            input_file: Path to pipe-delimited file
            output_file: Path for output CSV (optional)

        Returns:
            Path to converted CSV file
        """
        try:
            logger.info(f"Converting pipe-delimited file: {input_file}")

            if output_file is None:
                base_name = Path(input_file).stem
                output_file = self.output_dir / f"{base_name}_converted.csv"

            # Read pipe-delimited file
            df = pd.read_csv(input_file, sep='|', engine='python')

            # Save as comma-delimited with quotes where needed
            df.to_csv(output_file, index=False, quoting=csv.QUOTE_NONNUMERIC)

            logger.info(f"File converted successfully to {output_file}")
            self.converted_file = str(output_file)
            return str(output_file)
        except Exception as e:
            logger.error(f"Conversion failed: {str(e)}")
            raise

    def wait_for_processing(self, ftp: ftplib.FTP, expected_filename: str,
                           max_wait_minutes: int = 30, check_interval: int = 60):
        """
        Wait for FTP to process file and make it available

        Args:
            ftp: FTP connection object
            expected_filename: Name of file to wait for
            max_wait_minutes: Maximum time to wait in minutes
            check_interval: Time between checks in seconds
        """
        logger.info(f"Waiting for processing to complete. Looking for: {expected_filename}")

        max_attempts = (max_wait_minutes * 60) // check_interval
        attempts = 0

        while attempts < max_attempts:
            try:
                # List files in directory
                files = ftp.nlst()

                if expected_filename in files:
                    logger.info(f"File {expected_filename} found! Processing complete.")
                    return True

                attempts += 1
                logger.info(f"File not ready yet. Attempt {attempts}/{max_attempts}. Waiting {check_interval} seconds...")
                time.sleep(check_interval)

            except Exception as e:
                logger.warning(f"Error checking for file: {str(e)}")
                attempts += 1
                time.sleep(check_interval)

        logger.error(f"File {expected_filename} not found after {max_wait_minutes} minutes")
        raise TimeoutError(f"Processing timeout: File {expected_filename} not available")

    def analyze_email_changes(self, original_file: str, appended_file: str) -> Dict:
        """
        Analyze changes between original and appended files

        Args:
            original_file: Path to original file
            appended_file: Path to file after email appending

        Returns:
            Dictionary with analysis results
        """
        try:
            logger.info("Analyzing email changes...")

            # Read both files
            df_original = pd.read_csv(original_file)
            df_appended = pd.read_csv(appended_file)

            # Check if Email column exists in appended file
            email_col = None
            for col in df_appended.columns:
                if 'email' in col.lower():
                    email_col = col
                    break

            if email_col is None:
                logger.warning("No email column found in appended file")
                # Count rows instead
                self.emails_before = 0  # Original file has no emails
                self.emails_after = len(df_appended[df_appended[email_col].notna()]) if email_col else 0
            else:
                # Count non-empty emails
                emails_before_count = len(df_original[df_original[email_col].notna()]) if email_col in df_original.columns else 0
                emails_after_count = len(df_appended[df_appended[email_col].notna()])

                self.emails_before = emails_before_count
                self.emails_after = emails_after_count

                # Find new emails
                if email_col in df_original.columns:
                    original_emails = set(df_original[df_original[email_col].notna()][email_col].astype(str))
                else:
                    original_emails = set()

                appended_emails = set(df_appended[df_appended[email_col].notna()][email_col].astype(str))

                new_emails_set = appended_emails - original_emails
                self.new_emails = sorted(list(new_emails_set))

                # Check for duplicates in appended file
                email_counts = df_appended[email_col].value_counts()
                duplicates = email_counts[email_counts > 1].index.tolist()
                self.duplicate_emails = [email for email in duplicates if pd.notna(email)]

            results = {
                'total_records_before': len(df_original),
                'total_records_after': len(df_appended),
                'emails_before': self.emails_before,
                'emails_after': self.emails_after,
                'new_emails_count': len(self.new_emails),
                'duplicate_count': len(self.duplicate_emails),
                'new_emails': self.new_emails[:100],  # Limit to first 100 for display
                'duplicate_emails': self.duplicate_emails[:100]
            }

            logger.info(f"Analysis complete: {results['new_emails_count']} new emails added")
            return results

        except Exception as e:
            logger.error(f"Analysis failed: {str(e)}")
            raise

    def generate_excel_report(self, analysis_results: Dict, output_filename: str = None) -> str:
        """
        Generate detailed Excel report for client

        Args:
            analysis_results: Dictionary with analysis results
            output_filename: Path for output Excel file (optional)

        Returns:
            Path to generated Excel report
        """
        try:
            logger.info("Generating Excel report...")

            if output_filename is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_filename = self.output_dir / f"Email_Append_Report_{timestamp}.xlsx"

            wb = Workbook()

            # Summary Sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)

            # Title
            ws_summary['A1'] = "Email Append Report - Fushia FPN"
            ws_summary['A1'].font = Font(bold=True, size=14)
            ws_summary.merge_cells('A1:B1')

            ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws_summary.merge_cells('A2:B2')
            ws_summary['A3'] = ""

            # Summary data
            summary_data = [
                ['Metric', 'Value'],
                ['Total Records Before', analysis_results['total_records_before']],
                ['Total Records After', analysis_results['total_records_after']],
                ['Emails Before Append', analysis_results['emails_before']],
                ['Emails After Append', analysis_results['emails_after']],
                ['New Emails Added', analysis_results['new_emails_count']],
                ['Duplicate Emails Found', analysis_results['duplicate_count']],
            ]

            for row_idx, row_data in enumerate(summary_data, start=4):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 4:  # Header row
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')

            # Adjust column widths
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 20

            # New Emails Sheet
            if analysis_results['new_emails']:
                ws_new = wb.create_sheet("New Emails")
                ws_new['A1'] = "Newly Added Email Addresses"
                ws_new['A1'].font = Font(bold=True, size=12)
                ws_new['A1'].fill = header_fill
                ws_new['A1'].font = header_font

                for idx, email in enumerate(analysis_results['new_emails'], start=2):
                    ws_new.cell(row=idx, column=1, value=email)

                ws_new.column_dimensions['A'].width = 40

            # Duplicates Sheet
            if analysis_results['duplicate_emails']:
                ws_dup = wb.create_sheet("Duplicates")
                ws_dup['A1'] = "Duplicate Email Addresses"
                ws_dup['A1'].font = Font(bold=True, size=12)
                ws_dup['A1'].fill = header_fill
                ws_dup['A1'].font = header_font

                for idx, email in enumerate(analysis_results['duplicate_emails'], start=2):
                    ws_dup.cell(row=idx, column=1, value=email)

                ws_dup.column_dimensions['A'].width = 40

            # Save workbook
            wb.save(output_filename)
            logger.info(f"Excel report generated: {output_filename}")
            return str(output_filename)

        except Exception as e:
            logger.error(f"Report generation failed: {str(e)}")
            raise

    def _send_workflow_notifications(self, workflow_results: Dict):
        """
        Send email and Slack notifications after workflow completes

        Args:
            workflow_results: Dictionary with workflow results
        """
        if workflow_results['total_files'] == 0:
            return

        # Calculate totals
        total_files = workflow_results['total_files']
        successful = workflow_results['successful_files']
        failed = workflow_results['failed_files']
        total_new_emails = sum(
            r.get('analysis', {}).get('new_emails_count', 0)
            for r in workflow_results['files_processed']
            if r.get('success', False)
        )

        # Build notification message
        status = "SUCCESS" if failed == 0 else "PARTIAL SUCCESS" if successful > 0 else "FAILED"
        status_emoji = ":white_check_mark:" if failed == 0 else ":warning:" if successful > 0 else ":x:"

        # Slack message (short format)
        slack_message = f"*FPN Email Append Processing Complete*\n\n"
        slack_message += f"Status: {status_emoji} {status}\n"
        slack_message += f"Total Files: {total_files}\n"
        slack_message += f"Successful: {successful}\n"
        slack_message += f"Failed: {failed}\n"
        slack_message += f"New Emails Added: {total_new_emails}\n"

        if successful > 0:
            slack_message += f"\n*Successfully Processed:*\n"
            for result in workflow_results['files_processed']:
                if result['success']:
                    emails = result.get('analysis', {}).get('new_emails_count', 0)
                    slack_message += f"  • {result['filename']} ({emails} new emails)\n"

        if failed > 0:
            slack_message += f"\n*Failed Files:*\n"
            for result in workflow_results['files_processed']:
                if not result['success']:
                    slack_message += f"  • {result['filename']}: {result.get('error', 'Unknown error')}\n"

        # Email message (HTML format with more details)
        email_body = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                .header {{ background-color: {"#28a745" if failed == 0 else "#ffc107" if successful > 0 else "#dc3545"}; color: white; padding: 15px; }}
                .content {{ padding: 20px; }}
                .summary {{ background-color: #f8f9fa; padding: 15px; border-left: 4px solid {"#28a745" if failed == 0 else "#ffc107" if successful > 0 else "#dc3545"}; margin: 15px 0; }}
                .file-list {{ margin: 10px 0; }}
                .success {{ color: #28a745; }}
                .failure {{ color: #dc3545; }}
                table {{ border-collapse: collapse; width: 100%; margin: 15px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h2>FPN Email Append Processing Complete</h2>
            </div>
            <div class="content">
                <div class="summary">
                    <h3>Summary</h3>
                    <p><strong>Status:</strong> {status}</p>
                    <p><strong>Total Files Processed:</strong> {total_files}</p>
                    <p><strong>Successful:</strong> <span class="success">{successful}</span></p>
                    <p><strong>Failed:</strong> <span class="failure">{failed}</span></p>
                    <p><strong>Total New Emails Added:</strong> {total_new_emails}</p>
                    <p><strong>Completed:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
        """

        if successful > 0:
            email_body += """
                <h3>Successfully Processed Files</h3>
                <table>
                    <tr>
                        <th>Filename</th>
                        <th>New Emails</th>
                    </tr>
            """
            for result in workflow_results['files_processed']:
                if result['success']:
                    emails = result.get('analysis', {}).get('new_emails_count', 0)
                    email_body += f"""
                    <tr>
                        <td>{result['filename']}</td>
                        <td class="success">{emails}</td>
                    </tr>
                    """
            email_body += "</table>"

        if failed > 0:
            email_body += """
                <h3>Failed Files</h3>
                <table>
                    <tr>
                        <th>Filename</th>
                        <th>Error</th>
                    </tr>
            """
            for result in workflow_results['files_processed']:
                if not result['success']:
                    email_body += f"""
                    <tr>
                        <td>{result['filename']}</td>
                        <td class="failure">{result.get('error', 'Unknown error')}</td>
                    </tr>
                    """
            email_body += "</table>"

        email_body += """
            </div>
        </body>
        </html>
        """

        # Send notifications
        self.send_slack_notification(slack_message)
        self.send_email_notification(
            subject=f"FPN Email Append: {status} - {total_files} files processed",
            body=email_body,
            html=True
        )

    def _send_failure_notification(self, error_message: str):
        """
        Send notification when workflow fails

        Args:
            error_message: Error message
        """
        # Slack notification
        slack_message = f":x: *FPN Email Append Processing FAILED*\n\n"
        slack_message += f"Error: {error_message}\n"
        slack_message += f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Email notification
        email_body = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                .header {{ background-color: #dc3545; color: white; padding: 15px; }}
                .content {{ padding: 20px; }}
                .error {{ background-color: #f8d7da; padding: 15px; border-left: 4px solid #dc3545; margin: 15px 0; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h2>FPN Email Append Processing FAILED</h2>
            </div>
            <div class="content">
                <div class="error">
                    <h3>Error Details</h3>
                    <p>{error_message}</p>
                    <p><strong>Time:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
                <p>Please check the logs for more information.</p>
            </div>
        </body>
        </html>
        """

        self.send_slack_notification(slack_message)
        self.send_email_notification(
            subject="FPN Email Append: WORKFLOW FAILED",
            body=email_body,
            html=True
        )

    def process_single_file(self, ftp1: ftplib.FTP, sftp2: paramiko.SFTPClient,
                           filename: str, file_index: int, total_files: int) -> Dict:
        """
        Process a single file through the complete workflow

        Args:
            ftp1: FPN FTP connection
            sftp2: TowerData SFTP connection
            filename: Name of file to process
            file_index: Index of current file (for logging)
            total_files: Total number of files being processed

        Returns:
            Dictionary with processing results for this file
        """
        results = {
            'filename': filename,
            'success': False,
            'original_file': None,
            'converted_file': None,
            'appended_file': None,
            'report_file': None,
            'error': None
        }

        try:
            logger.info("\n" + "="*70)
            logger.info(f"Processing file {file_index}/{total_files}: {filename}")
            logger.info("="*70)

            # Step 1: Download file from FPN
            logger.info(f"=== Step 1: Downloading {filename} from FPN ===")
            source_dir = self.config['source_filename']
            remote_filepath = f"{source_dir}{filename}"

            original_file = self.download_from_ftp(ftp1, remote_filepath, filename)
            results['original_file'] = original_file

            # Step 2: Move file to previous/ folder on FPN
            logger.info(f"=== Step 2: Moving {filename} to previous/ folder ===")
            previous_dir = self.config.get('fpn_previous_dir', f"{source_dir}previous/").rstrip('/')
            # Ensure the previous/ directory exists
            self.ensure_ftp_directory(ftp1, previous_dir)
            dest_filepath = f"{previous_dir}/{filename}"
            self.move_ftp_file(ftp1, remote_filepath, dest_filepath)

            # Step 3: Convert pipe to CSV
            logger.info("=== Step 3: Converting file format ===")
            converted_file = self.convert_pipe_to_csv(original_file)
            results['converted_file'] = converted_file

            # Step 4: Upload to TowerData SFTP
            logger.info("=== Step 4: Uploading to TowerData for email append ===")
            upload_filename = f"{Path(filename).stem}_converted.csv"
            self.upload_to_sftp(sftp2, converted_file, f"uploads_hh/{upload_filename}")

            # Step 5: Wait for processing and download result
            logger.info("=== Step 5: Waiting for TowerData processing ===")
            processed_file_path = self.wait_for_processing_sftp(
                sftp2,
                upload_filename,
                max_wait_minutes=self.config.get('max_wait_minutes', 30),
                check_interval=self.config.get('check_interval', 60)
            )

            # Step 6: Download processed file
            logger.info("=== Step 6: Downloading processed file ===")
            result_filename = Path(processed_file_path).name
            appended_file = self.download_from_sftp(
                sftp2,
                processed_file_path,
                result_filename
            )
            results['appended_file'] = appended_file

            # Step 7: Upload result back to FPN
            logger.info("=== Step 7: Uploading result to FPN ===")

            # Reconnect to FTP if connection is stale (after long TowerData wait)
            try:
                # Test if connection is still alive
                ftp1.pwd()
            except Exception as e:
                logger.warning(f"FTP connection appears stale, reconnecting: {str(e)}")
                try:
                    ftp1.quit()
                except:
                    pass
                # Reconnect
                ftp1 = self.connect_ftp(
                    self.config['ftp1_server'],
                    self.config['ftp1_username'],
                    self.config['ftp1_password'],
                    self.config.get('ftp1_port', 21)
                )
                logger.info("[OK] FTP reconnection successful")

            fpn_dest_dir = self.config.get('fpn_destination_dir', '/PreScreen Delivery/Fushia/to_FPN/').rstrip('/')
            # Ensure the destination directory exists
            self.ensure_ftp_directory(ftp1, fpn_dest_dir)
            fpn_result_path = f"{fpn_dest_dir}/{result_filename}"
            self.upload_to_ftp(ftp1, appended_file, fpn_result_path)

            # Step 8: Analyze changes
            logger.info("=== Step 8: Analyzing email changes ===")
            analysis_results = self.analyze_email_changes(converted_file, appended_file)

            # Step 9: Generate Excel report
            logger.info("=== Step 9: Generating Excel report ===")
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_filename = f"Email_Append_Report_{Path(filename).stem}_{timestamp}.xlsx"
            report_file = self.generate_excel_report(analysis_results,
                                                    self.output_dir / report_filename)
            results['report_file'] = report_file
            results['analysis'] = analysis_results
            results['success'] = True

            logger.info(f"[OK] File {filename} processed successfully!")
            logger.info(f"  New emails added: {analysis_results['new_emails_count']}")
            logger.info(f"  Result uploaded to: {fpn_result_path}")

        except Exception as e:
            logger.error(f"[FAIL] Failed to process {filename}: {str(e)}")
            results['error'] = str(e)

        return results

    def process_workflow(self) -> Dict:
        """
        Execute complete workflow for all files with keyword filter

        Returns:
            Dictionary with workflow results
        """
        workflow_results = {
            'success': False,
            'files_processed': [],
            'total_files': 0,
            'successful_files': 0,
            'failed_files': 0,
            'error': None
        }

        ftp1 = None
        sftp2 = None
        ssh2 = None

        try:
            logger.info("\n" + "="*70)
            logger.info("STARTING BATCH EMAIL APPEND WORKFLOW")
            logger.info("="*70)

            # Step 1: Connect to FPN FTP
            logger.info("\n=== Connecting to FPN FTP Server ===")
            if SKIP_FTP1_DOWNLOAD:
                logger.info("[WARNING] SKIP_FTP1_DOWNLOAD is enabled - using test mode")
                logger.info("Processing single test file only")

                # Process single test file
                original_file = LOCAL_TEST_FILE
                if not os.path.exists(original_file):
                    raise FileNotFoundError(f"Local test file not found: {original_file}")

                # Connect to TowerData
                logger.info("\n=== Connecting to TowerData SFTP Server ===")
                sftp2, ssh2 = self.connect_sftp(
                    self.config['ftp2_server'],
                    self.config['ftp2_username'],
                    self.config['ftp2_password'],
                    self.config.get('ftp2_port', 22)
                )

                # Create a mock result for test file
                test_result = {
                    'filename': Path(original_file).name,
                    'success': False,
                    'original_file': original_file,
                    'converted_file': None,
                    'appended_file': None,
                    'report_file': None,
                    'error': None
                }

                try:
                    # Convert and process
                    logger.info("=== Converting file format ===")
                    converted_file = self.convert_pipe_to_csv(original_file)
                    test_result['converted_file'] = converted_file

                    logger.info("=== Uploading to TowerData ===")
                    upload_filename = self.config.get('upload_filename', Path(converted_file).name)
                    self.upload_to_sftp(sftp2, converted_file, f"uploads_hh/{upload_filename}")

                    logger.info("=== Waiting for processing ===")
                    processed_file_path = self.wait_for_processing_sftp(
                        sftp2, upload_filename,
                        max_wait_minutes=self.config.get('max_wait_minutes', 30),
                        check_interval=self.config.get('check_interval', 60)
                    )

                    logger.info("=== Downloading result ===")
                    result_filename = Path(processed_file_path).name
                    appended_file = self.download_from_sftp(sftp2, processed_file_path, result_filename)
                    test_result['appended_file'] = appended_file

                    logger.info("=== Analyzing changes ===")
                    analysis_results = self.analyze_email_changes(converted_file, appended_file)

                    logger.info("=== Generating report ===")
                    report_file = self.generate_excel_report(analysis_results)
                    test_result['report_file'] = report_file
                    test_result['analysis'] = analysis_results
                    test_result['success'] = True

                    workflow_results['successful_files'] = 1
                    logger.info(f"\n[OK] Test file processed successfully!")
                    logger.info(f"  New emails added: {analysis_results['new_emails_count']}")

                except Exception as e:
                    test_result['error'] = str(e)
                    workflow_results['failed_files'] = 1
                    logger.error(f"[FAIL] Test file processing failed: {str(e)}")

                workflow_results['files_processed'].append(test_result)
                workflow_results['total_files'] = 1
                workflow_results['success'] = test_result['success']

            else:
                # Production mode: Process all files from FPN
                ftp1 = self.connect_ftp(
                    self.config['ftp1_server'],
                    self.config['ftp1_username'],
                    self.config['ftp1_password'],
                    self.config.get('ftp1_port', 21)
                )

                # Step 2: List files with keyword filter
                logger.info("\n=== Listing files from FPN ===")
                source_dir = self.config['source_filename']
                keyword = self.config.get('file_filter_keyword', 'email')
                files_to_process = self.list_ftp_files(ftp1, source_dir, keyword)
                if files_to_process:
                    # Send immediate notification that files were detected
                    file_list = ", ".join(files_to_process)

                    # Slack notification (remove emoji for Windows compatibility)
                    self.send_slack_notification(
                        message=f"New file(s) detected in {source_dir}:\n{file_list}\n\nProcessing will begin shortly...",
                        title="Files Detected in FPN Folder"
                    )

                    # Email notification
                    email_body = f"""
                    <html>
                    <head>
                        <style>
                            body {{ font-family: Arial, sans-serif; }}
                            .header {{ background-color: #007bff; color: white; padding: 15px; }}
                            .content {{ padding: 20px; }}
                            .file-list {{ background-color: #f8f9fa; padding: 15px; border-left: 4px solid #007bff; margin: 15px 0; }}
                            .info {{ color: #0056b3; }}
                        </style>
                    </head>
                    <body>
                        <div class="header">
                            <h2>New Files Detected in FPN Folder</h2>
                        </div>
                        <div class="content">
                            <div class="file-list">
                                <h3>Detected Files:</h3>
                                <ul>
                    """

                    for file in files_to_process:
                        email_body += f"<li>{file}</li>\n"

                    email_body += f"""
                                </ul>
                            </div>
                            <p><strong>Location:</strong> {source_dir}</p>
                            <p><strong>File Count:</strong> {len(files_to_process)}</p>
                            <p><strong>Detection Time:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                            <p class="info"><strong>Status:</strong> Processing will begin shortly. You will receive another notification when processing completes.</p>
                        </div>
                    </body>
                    </html>
                    """

                    self.send_email_notification(
                        subject=f"FPN Alert: {len(files_to_process)} new file(s) detected",
                        body=email_body,
                        html=True
                    )

                    logger.info(f"[OK] File detection notifications sent for {len(files_to_process)} file(s)")

                if not files_to_process:
                    logger.warning(f"No files found with keyword '{keyword}' in {source_dir}")
                    workflow_results['success'] = True
                    workflow_results['total_files'] = 0
                    return workflow_results

                workflow_results['total_files'] = len(files_to_process)
                logger.info(f"\nFound {len(files_to_process)} files to process:")
                for idx, fname in enumerate(files_to_process, 1):
                    logger.info(f"  {idx}. {fname}")

                # Step 3: Connect to TowerData
                logger.info("\n=== Connecting to TowerData SFTP Server ===")
                sftp2, ssh2 = self.connect_sftp(
                    self.config['ftp2_server'],
                    self.config['ftp2_username'],
                    self.config['ftp2_password'],
                    self.config.get('ftp2_port', 22)
                )

                # Step 4: Process each file
                for idx, filename in enumerate(files_to_process, 1):
                    result = self.process_single_file(ftp1, sftp2, filename, idx, len(files_to_process))
                    workflow_results['files_processed'].append(result)

                    if result['success']:
                        workflow_results['successful_files'] += 1
                    else:
                        workflow_results['failed_files'] += 1

                workflow_results['success'] = workflow_results['failed_files'] == 0

            # Final summary
            logger.info("\n" + "="*70)
            logger.info("BATCH WORKFLOW SUMMARY")
            logger.info("="*70)
            logger.info(f"Total files: {workflow_results['total_files']}")
            logger.info(f"Successful: {workflow_results['successful_files']}")
            logger.info(f"Failed: {workflow_results['failed_files']}")

            if workflow_results['successful_files'] > 0:
                logger.info("\nSuccessfully processed files:")
                for result in workflow_results['files_processed']:
                    if result['success']:
                        logger.info(f"  [OK] {result['filename']}")
                        if 'analysis' in result:
                            logger.info(f"    - New emails: {result['analysis']['new_emails_count']}")

            if workflow_results['failed_files'] > 0:
                logger.info("\nFailed files:")
                for result in workflow_results['files_processed']:
                    if not result['success']:
                        logger.info(f"  [FAIL] {result['filename']}: {result.get('error', 'Unknown error')}")

            # Send notifications after workflow completes
            self._send_workflow_notifications(workflow_results)

        except Exception as e:
            logger.error(f"\nWorkflow failed: {str(e)}")
            workflow_results['error'] = str(e)
            import traceback
            traceback.print_exc()

            # Send failure notification
            self._send_failure_notification(str(e))

        finally:
            # Close FTP/SFTP connections
            if ftp1:
                try:
                    ftp1.quit()
                    logger.info("FPN FTP connection closed")
                except:
                    pass
            if sftp2:
                try:
                    sftp2.close()
                    logger.info("TowerData SFTP connection closed")
                except:
                    pass
            if ssh2:
                try:
                    ssh2.close()
                except:
                    pass

        return workflow_results


def main():
    """
    Main entry point for standalone execution
    Load configuration and run processor
    """
    # Import configuration
    try:
        from config import FTP_CONFIG
    except ImportError:
        logger.error("config.py not found. Please create configuration file.")
        logger.info("See config_template.py for an example")
        return

    # Create processor and run workflow
    processor = FTPEmailProcessor(FTP_CONFIG)
    results = processor.process_workflow()

    if results['success']:
        logger.info("\n" + "="*70)
        logger.info("ALL FILES PROCESSED SUCCESSFULLY")
        logger.info("="*70)
        logger.info(f"\nTotal files processed: {results['total_files']}")
        logger.info(f"Successful: {results['successful_files']}")
        logger.info(f"Failed: {results['failed_files']}")

        if results['successful_files'] > 0:
            total_new_emails = sum(
                r.get('analysis', {}).get('new_emails_count', 0)
                for r in results['files_processed']
                if r['success']
            )
            logger.info(f"\nTotal new emails added across all files: {total_new_emails}")
    else:
        logger.error(f"\nWORKFLOW FAILED")
        if results.get('error'):
            logger.error(f"Error: {results['error']}")
        if results['failed_files'] > 0:
            logger.error(f"\nFailed files: {results['failed_files']}/{results['total_files']}")


if __name__ == "__main__":
    main()
